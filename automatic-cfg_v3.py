import os
# import xlsxwriter
from openpyxl import load_workbook

def read_input(doc):
	with open(doc) as input_list:
		input_list = input_list.readlines()
		# j'enlève le saut de page
		input_list[1] = input_list[1].rstrip("\n")
		adresses_list = input_list[1].split(" ")
		input_list[4] = input_list[4].rstrip("\n")
		values_list = input_list[4].split(" ")      
	return adresses_list, values_list

# création de l'excel avec la mise en forme spécifique
def excel_creation(general_path):
	# se place au chemin qu'on lui donne 
	excel_directory_move = os.chdir(general_path)
	# 
	wb = xlsxwriter.Workbook('cfg_input.xlsx')
	worksheet = wb.add_worksheet()
	wrap_format = wb.add_format({'text_wrap': True}) # permet de sauter des lignes dans une cellule
	cell_format_title = wb.add_format({'bold': True})
	cell_format_title.set_font_size(13)
	cell_format_title.set_align('center')
	cell_format_title.set_align('vcenter')
	worksheet.set_column('A:A', 40)
	worksheet.write(0,0,"Adress", cell_format_title)
	worksheet.set_column('B:B', 40)
	worksheet.write(0,1,"Values", cell_format_title)
	return wb, worksheet, wrap_format

def writing(value1, value2, wb, worksheet, wrap_format):
	cell_format_folder = wb.add_format({})
	cell_format_folder.set_align('vcenter')
	cell_format_folder.set_align('center')
	cell_format = wb.add_format({})
	x=1
	for i, j in zip(value1, value2):
		worksheet.write(x, 0, i)
		worksheet.write(x, 1, j, wrap_format)
		x += 1
	wb.close()

def reading(file):
	adresses_list = []
	values_list = []
	wb = load_workbook(file)
	# sheet = "\"{}\"".format(wb.sheetnames[0]) pour rappel de comment ajouter des éléments à une chaine de caractère
	sheet = wb.sheetnames[0]
	active_ws=wb[sheet]
	row_rangeA = active_ws['A2':'A83']
	row_rangeF = active_ws['F2':'F83']
	# for row in active_ws.iter_rows(min_row=2,max_row=83,max_col=1):
	# for row in row_range:
		# print(row)
	columnName = active_ws.cell(row=1, column=6)
	name = columnName.value
	print(name)
	for rowA, rowF in zip(row_rangeA, row_rangeF):
		# print(rowA, rowG)
		for cellA, cellF in zip(rowA, rowF):
			if cellA.value == None and cellF.value == None:
				continue
			else :
				if "(" and ")" in cellA.value:
					if cellF.value == None:
						continue
					else :
						print(cellA.value[0:4])
						adresses_list.append(cellA.value[0:4])
						values_list.append(cellF.value)
				elif cellF.value == None:
					continue
				elif cellF.value == "-":
					continue
				else :
					print(cellA.value, cellF.value)
					adresses_list.append(cellA.value)
					values_list.append(cellF.value)
	return adresses_list, values_list, name

def system_lines(list1, list2, column):
	newList1 = []
	newList2 = []
	tab = "\t\t\t\t\t\t\t\t\t\t"
	lenMax = len(list1)
	i = 0
	if 40 <= lenMax < 47 :
		while i < lenMax:
			if i <= 36 :
				newList1.append(' '.join(list1[i:i+8]))
				newList2.append(' '.join(list2[i:i+8]))
			else :
				newList1.append(' '.join(list1[i:lenMax]))
				newList2.append(' '.join(list2[i:lenMax]))
			i += 8
	# elif 32 <= lenMax < 40 :
	adresses = "\t\t\t\tsn65dsi84,addresses = < {}\n{}{}\n{}{}\n{}{}\n{}{}\n{}{}>;".format(newList1[0], tab, newList1[1], tab, newList1[2], tab, newList1[3], tab, newList1[4], tab, newList1[5])
	values = "\t\t\t\tsn65dsi84,values =    < {}\n{}{}\n{}{}\n{}{}\n{}{}\n{}{}>;".format(newList2[0], tab, newList2[1], tab, newList2[2], tab, newList2[3], tab, newList2[4], tab, newList2[5])
	print("\n\n{}".format(column), "\n\n", adresses, "\n\n", values)

def main():
	# on obtient le chemin d'où le code est lancé
	general_path = os.path.abspath(os.getcwd())
	input_document ="SN65DSI84 - register map.xlsx"
	adresses_list, values_list, nameColumn = reading(input_document)
	system_lines(adresses_list, values_list, nameColumn)

if __name__ == "__main__":
	main()