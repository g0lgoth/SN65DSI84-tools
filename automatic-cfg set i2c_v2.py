import os
# import xlsxwriter
from openpyxl import load_workbook
import tkinter
# from subprocess import Popen
import subprocess
from pathlib import Path

def cfg_creation(path):
	file_name = "cfg.txt"
	# file_path = os.chdir(path)
	file_path = Path("{}\\{}".format(path, file_name))
	print(file_path)
	if file_path.is_file() == True :
		return
	else :
		cfg_file = open("cfg.txt", "w")

def write_on_file(file, start_line, ):
	cfg_file = open(file, "w")
	cfg_file = open(file, "w")
	cfg_file = close(file)

general_path = os.path.abspath(os.getcwd())
cfg_creation(general_path)

# def read_input(doc):
# 	with open(doc) as input_list:
# 		input_list = input_list.readlines()
# 		# j'enlève le saut de page
# 		input_list[1] = input_list[1].rstrip("\n")
# 		adresses_list = input_list[1].split(" ")
# 		input_list[4] = input_list[4].rstrip("\n")
# 		values_list = input_list[4].split(" ")      
# 	return adresses_list, values_list

# # création de l'excel avec la mise en forme spécifique
# def excel_creation(general_path):
# 	# se place au chemin qu'on lui donne 
# 	excel_directory_move = os.chdir(general_path)
# 	# 
# 	wb = xlsxwriter.Workbook('cfg_input.xlsx')
# 	worksheet = wb.add_worksheet()
# 	wrap_format = wb.add_format({'text_wrap': True}) # permet de sauter des lignes dans une cellule
# 	cell_format_title = wb.add_format({'bold': True})
# 	cell_format_title.set_font_size(13)
# 	cell_format_title.set_align('center')
# 	cell_format_title.set_align('vcenter')
# 	worksheet.set_column('A:A', 40)
# 	worksheet.write(0,0,"Adress", cell_format_title)
# 	worksheet.set_column('B:B', 40)
# 	worksheet.write(0,1,"Values", cell_format_title)
# 	return wb, worksheet, wrap_format

# def writing(value1, value2, wb, worksheet, wrap_format):
# 	cell_format_folder = wb.add_format({})
# 	cell_format_folder.set_align('vcenter')
# 	cell_format_folder.set_align('center')
# 	cell_format = wb.add_format({})
# 	x=1
# 	for i, j in zip(value1, value2):
# 		worksheet.write(x, 0, i)
# 		worksheet.write(x, 1, j, wrap_format)
# 		x += 1
# 	wb.close()

# def reading(file, column_valueA, column_valueB):
# 	adresses_list = []
# 	values_list = []
# 	wb = load_workbook(file)
# 	# sheet = "\"{}\"".format(wb.sheetnames[0]) pour rappel de comment ajouter des éléments à une chaine de caractère
# 	sheet = wb.sheetnames[0]
# 	active_ws=wb[sheet]
# 	row_range1 = active_ws['{}1'.format(column_valueA):'{}100'.format(column_valueA)]
# 	row_range2 = active_ws['{}1'.format(column_valueB):'{}100'.format(column_valueB)]
# 	# for row in active_ws.iter_rows(min_row=2,max_row=83,max_col=1):
# 	# for row in row_range:
# 		# print(row)
# 	columnName = active_ws.cell(row=1, column=6)
# 	name = columnName.value
# 	print(name)
# 	for rowA, rowB in zip(row_range1, row_range2):
# 		# print(rowA, rowG)
# 		for cellA, cellB in zip(rowA, rowB):
# 			# if cellA.value == None and cellF.value == None:
# 			if cellA.value == None:
# 				continue
# 			elif "0x" not in cellA.value:
# 			 	continue
# 			else :
# 				if "(" and ")" in cellA.value:
# 					if cellB.value == None:
# 						continue
# 					else :
# 						print(cellA.value[0:4])
# 						adresses_list.append(cellA.value[0:4])
# 						values_list.append(cellB.value)
# 				elif cellB.value == None:
# 					continue
# 				elif cellB.value == "-":
# 					continue

# 				else :
# 					print(cellA.value, cellB.value)
# 					adresses_list.append(cellA.value)
# 					values_list.append(cellB.value)
# 	return adresses_list, values_list, name

# def i2cCommand(i2c_number, driver_adr, register_adr, register_val, local_path):
# 	sh_directory_move = os.chdir(local_path)
# 	with open('sn65dsi84.sh', 'w', newline = '\n') as sh_file:
# 		sh_file.write('#!/bin/sh\n')
# 		sh_directory_move = os.chdir(local_path)
# 		for i, j in zip(register_adr, register_val):
# 			sh_file.write('i2cset -y {} {} {} {}\n'.format(i2c_number, driver_adr, i, j))
# 	sh_file.close()

# def batchCommand(adb_path, local_path):
# 	batch_directory_move = os.chdir(local_path)
# 	with open('sn65dsi84.bat', 'w') as batch:
# 		batch.write('cd "{}"\n'.format(adb_path))
# 		# batch.write('adb root\n')
# 		# batch.write('adb disable-verity\n')
# 		# batch.write('adb reboot\n')
# 		# batch.write('adb wait-for-device\n')
# 		# batch.write('adb root\n')
# 		# batch.write('adb wait-for-device\n')
# 		# batch.write('adb remount\n')
# 		# batch.write('adb wait-for-device\n')
# 		# batch.write('TIMEOUT 5\n')
# 		batch.write('adb push "C:\\platform-tools_r28.0.1-windows\\script gogo\\sn65dsi84.sh" /system/bin\n')
# 		batch.write('adb wait-for-device\n')
# 		batch.write('adb shell chmod +x /system/bin/sn65dsi84.sh\n')
# 		batch.write('adb wait-for-device\n')
# 		batch.write('adb shell ./system/bin/sn65dsi84.sh\n')
# 		batch.close()
# 	# procId = subprocess.Popen('adb shell', stdin = subprocess.PIPE)
# 	# procId.communicate('./system/bin/sn65dsi84.sh')
# 	subprocess.call([r'{}sn65dsi84.bat'.format(local_path)])

# def system_lines(list1, list2, column):
# 	newList1 = []
# 	newList2 = []
# 	tab = "\t\t\t\t\t\t\t\t\t\t"
# 	lenMax = len(list1)
# 	print(lenMax)
# 	i = 0
# 	if 40 <= lenMax < 47 :
# 		while i < lenMax:
# 			if i <= 36 :
# 				newList1.append(' '.join(list1[i:i+8]))
# 				newList2.append(' '.join(list2[i:i+8]))
# 			else :
# 				newList1.append(' '.join(list1[i:lenMax]))
# 				newList2.append(' '.join(list2[i:lenMax]))
# 			i += 8
# 		adresses = "\t\t\t\tsn65dsi84,addresses = < {}\n{}{}\n{}{}\n{}{}\n{}{}\n{}{}>;".format(newList1[0], tab, newList1[1], tab, newList1[2], tab, newList1[3], tab, newList1[4], tab, newList1[5])
# 		values = "\t\t\t\tsn65dsi84,values =    < {}\n{}{}\n{}{}\n{}{}\n{}{}\n{}{}>;".format(newList2[0], tab, newList2[1], tab, newList2[2], tab, newList2[3], tab, newList2[4], tab, newList2[5])
# 		print("\n\n{}".format(column), "\n\n", adresses, "\n\n", values)
# 	elif 32 <= lenMax < 39 :
# 		while i < lenMax:
# 			if i <= 28 :
# 				newList1.append(' '.join(list1[i:i+8]))
# 				newList2.append(' '.join(list2[i:i+8]))
# 			else :
# 				newList1.append(' '.join(list1[i:lenMax]))
# 				newList2.append(' '.join(list2[i:lenMax]))
# 			i += 8
# 		adresses = "\t\t\t\tsn65dsi84,addresses = < {}\n{}{}\n{}{}\n{}{}\n{}{}>;".format(newList1[0], tab, newList1[1], tab, newList1[2], tab, newList1[3], tab, newList1[4])
# 		values = "\t\t\t\tsn65dsi84,values =    < {}\n{}{}\n{}{}\n{}{}\n{}{}>;".format(newList2[0], tab, newList2[1], tab, newList2[2], tab, newList2[3], tab, newList2[4])
# 		print("\n\n{}".format(column), "\n\n", adresses, "\n\n", values)
# 	else :
# 		while i < lenMax:
# 			if i <= 22 :
# 				newList1.append(' '.join(list1[i:i+8]))
# 				newList2.append(' '.join(list2[i:i+8]))
# 			else :
# 				newList1.append(' '.join(list1[i:lenMax]))
# 				newList2.append(' '.join(list2[i:lenMax]))
# 			i += 8
# 		adresses = "\t\t\t\tsn65dsi84,addresses = < {}\n{}{}\n{}{}\n{}{}>;".format(newList1[0], tab, newList1[1], tab, newList1[2], tab, newList1[3])
# 		values = "\t\t\t\tsn65dsi84,values =    < {}\n{}{}\n{}{}\n{}{}>;".format(newList2[0], tab, newList2[1], tab, newList2[2], tab, newList2[3])
# 		print("\n\n{}".format(column), "\n\n", adresses, "\n\n", values)

# def main():
# 	# on obtient le chemin d'où le code est lancé
# 	general_path = os.path.abspath(os.getcwd())
# 	adb_path = "C:\\platform-tools_r28.0.1-windows\\"
# 	local_path = "C:\\platform-tools_r28.0.1-windows\\script gogo\\"
# 	i2c_number = 2
# 	driver_adr = "0x2c"
# 	input_document ="SN65DSI84 - register map.xlsx"
# 	column1 = 'A'
# 	column2 = 'B'
# 	adresses_list, values_list, nameColumn = reading(input_document, column1, column2)
# 	system_lines(adresses_list, values_list, nameColumn)
# 	i2cCommand(i2c_number, driver_adr ,adresses_list, values_list, local_path)
# 	batchCommand(adb_path, local_path)

# if __name__ == "__main__":
	# main()